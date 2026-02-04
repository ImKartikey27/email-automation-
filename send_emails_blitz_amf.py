"""
Email Sender Script - Production Cold Email Automation
======================================================
📧 Gmail API integration (safe from bans)
🔄 Duplicate prevention + Resume capability
💾 Status tracking in Excel
⏱️ Safe batch sending (30 emails/day)
🎯 Optimized for cold email campaigns

Author: Built for Kartikey Sangal
Date: 2025-01-14
Version: Production Ready (No AI)
"""

import os
import sys
import time
import random
from datetime import datetime
from openpyxl import load_workbook
from dotenv import load_dotenv
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from googleapiclient.discovery import build
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import base64
import pickle

# ==================== CONFIGURATION ====================

load_dotenv()

# Paths
EXCEL_DIR = "/Users/kartikeysangal/Documents/email-automation/"
EXCEL_FILENAME = "Companies-Noida.xlsx"
# EXCEL_FILENAME = "Squadstack_20260115_122139.xlsx"

EXCEL_FILE = os.path.join(EXCEL_DIR, EXCEL_FILENAME)

RESUME_PATH = "/Users/kartikeysangal/Documents/email-automation/KARTIKEY-Resume.pdf"
TOKEN_FILE = "/Users/kartikeysangal/Documents/email-automation/token.pickle"
CREDENTIALS_FILE = "/Users/kartikeysangal/Documents/email-automation/credentials.json"

# Sheet names
INPUT_SHEET = "found_blitz"

# Gmail API setup
SCOPES = ['https://www.googleapis.com/auth/gmail.send']

# Batch configuration - SAFE FOR PRIMARY ACCOUNT
BATCH_CONFIG = {
    'emails_per_day': 22,            # Safe limit for primary Gmail
    'delay_min': 20,                 # Minimum 20 seconds
    'delay_max': 35,                 # Maximum 35 seconds (random)
    'batch_size': 10,                # Send 10, then longer break
    'batch_break': 300,              # 5 minutes between batches
}

# Email template
EMAIL_SUBJECT = "Quick question about {company_name}'s backend"

EMAIL_BODY_TEMPLATE = """
Hi {first_name},<br><br>

I came across {company_name} and was curious, what does your backend stack look like these days?<br><br>

I'm a Full-Stack Engineer with backend focus. Recently, I've been building production systems for clients:<br><br>

<b>CFA Exam Platform</b> (for an EdTech startup)<br>
Built the complete backend infrastructure — NestJS, Redis, AWS Lambda, Postgres, Docker<br>
→ <a href="https://blue-nelumbo.vercel.app/">Live</a><br><br>

<b>Investor-Facing Website + CMS</b> (for a real estate client)<br>
Full-stack animated site with custom CMS — Next.js, Three.js, Strapi<br>
→ <a href="https://apartment-eleven-eleven-plum.vercel.app/">Live</a><br><br>

I've also handled smaller DevOps and deployment fixes for clients on Upwork.<br><br>

If you're scaling the backend team, I'd love to chat.<br><br>

Best,<br>
Kartikey Sangal<br>
<a href="https://drive.google.com/file/d/1G_JubtQoUVfkHtbtEvfu8g_0pZS78Jsa/view">Resume</a> | <a href="https://www.kartikeyx.me">Portfolio</a> | <a href="https://www.github.com/ImKartikey27">GitHub</a> | <a href="https://www.linkedin.com/in/kartikey-sangal-752567301/">LinkedIn</a>
"""

# ==================== GMAIL API SETUP ====================

def get_gmail_service():
    """Authenticate and return Gmail API service"""
    creds = None

    # Load existing token
    if os.path.exists(TOKEN_FILE):
        with open(TOKEN_FILE, 'rb') as token:
            creds = pickle.load(token)

    # If no valid credentials, authenticate
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            if not os.path.exists(CREDENTIALS_FILE):
                print(f"\n❌ ERROR: credentials.json not found at {CREDENTIALS_FILE}")
                sys.exit(1)

            flow = InstalledAppFlow.from_client_secrets_file(
                CREDENTIALS_FILE, SCOPES)
            creds = flow.run_local_server(port=0)

        # Save token
        with open(TOKEN_FILE, 'wb') as token:
            pickle.dump(creds, token)

    return build('gmail', 'v1', credentials=creds)

# ==================== EMAIL FUNCTIONS ====================

def create_email_with_attachment(to, subject, html_body, attachment_path=None):
    """Create email with HTML body and attachment"""
    message = MIMEMultipart("mixed")
    message["to"] = to
    message["subject"] = subject

    # Create alternative part (plain + html)
    alternative = MIMEMultipart("alternative")

    # Plain-text fallback
    plain_text = (
        "Hi,\n\n"
        "I hope you're doing well. I'm a Full-Stack Engineer reaching out about opportunities.\n\n"
        "Please view this email in an HTML-capable client to see full details.\n\n"
        "Best regards,\n"
        "Kartikey Sangal\n"
        "https://www.kartikeyx.me"
    )

    alternative.attach(MIMEText(plain_text, "plain"))
    alternative.attach(MIMEText(html_body, "html"))

    message.attach(alternative)

    # Attachment
# Attachment (ONLY if provided)
    if attachment_path and os.path.exists(attachment_path):
        with open(attachment_path, "rb") as f:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(f.read())

        encoders.encode_base64(part)
        part.add_header(
            "Content-Disposition",
            f'attachment; filename="{os.path.basename(attachment_path)}"'
        )
        message.attach(part)


    raw_message = base64.urlsafe_b64encode(message.as_bytes()).decode("utf-8")
    return {"raw": raw_message}


def send_email(service, to, subject, body, attachment_path=None):
    """Send email via Gmail API"""
    try:
        message = create_email_with_attachment(to, subject, body, attachment_path)
        sent_message = service.users().messages().send(
            userId='me',
            body=message
        ).execute()
        return {'success': True, 'message_id': sent_message['id']}
    except Exception as e:
        return {'success': False, 'error': str(e)}

# ==================== EXCEL OPERATIONS ====================

def read_and_filter_contacts():
    """Read contacts and filter already sent"""
    print("📊 Reading contacts from Excel...")
    
    if not os.path.exists(EXCEL_FILE):
        print(f"❌ ERROR: Excel file not found at {EXCEL_FILE}")
        sys.exit(1)

    wb = load_workbook(EXCEL_FILE)

    if INPUT_SHEET not in wb.sheetnames:
        print(f"❌ ERROR: Sheet '{INPUT_SHEET}' not found!")
        print(f"Available sheets: {wb.sheetnames}")
        sys.exit(1)

    ws = wb[INPUT_SHEET]
    headers = [cell.value for cell in ws[1]]

    # Check required columns
    required_cols = ['personal_email', 'first_name', 'company_linkedin_url', 'position']
    missing_cols = [col for col in required_cols if col not in headers]
    
    if missing_cols:
        print(f"❌ ERROR: Missing required columns: {missing_cols}")
        print(f"Available columns: {headers}")
        sys.exit(1)

    # Add email_status column if not exists
    if 'email_status' not in headers:
        ws.cell(1, len(headers) + 1, 'email_status')
        wb.save(EXCEL_FILE)
        headers.append('email_status')
        print("✅ Added 'email_status' column to Excel\n")

    # Get column indices
    email_idx = headers.index('personal_email')
    status_idx = headers.index('email_status')
    first_name_idx = headers.index('first_name')
    company_idx = headers.index('company_linkedin_url')
    position_idx = headers.index('position')
    
    # Check if company_name column exists (optional)
    company_name_idx = headers.index('company_name') if 'company_name' in headers else None

    # Read all contacts
    all_contacts = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        email = row[email_idx].value
        status = row[status_idx].value if status_idx < len(row) else None
        first_name = row[first_name_idx].value if first_name_idx < len(row) else ''
        company_linkedin = row[company_idx].value if company_idx < len(row) else ''
        position = row[position_idx].value if position_idx < len(row) else ''
        
        # Get company name from column if available
        company_name_col = row[company_name_idx].value if company_name_idx is not None and company_name_idx < len(row) else None

        if email and email.strip():
            all_contacts.append({
                'row_num': idx,
                'email': email.strip(),
                'first_name': first_name or 'there',
                'company_name': extract_company_name(company_linkedin, company_name_col),
                'position': position or 'team member',
                'status': status,
                'row_cells': row
            })

    wb.close()

    # Filter: Only contacts without status (not sent yet)
    pending = [c for c in all_contacts if not c['status'] or c['status'].strip() == '']
    already_sent = len(all_contacts) - len(pending)

    print(f"✅ Total contacts: {len(all_contacts)}")
    print(f"✅ Already sent: {already_sent}")
    print(f"✅ Pending: {len(pending)}\n")

    return pending

def extract_company_name(linkedin_url, company_name_column=None):
    """Extract company name from Company Name column or LinkedIn URL"""
    
    # First priority: Check if company name is directly available in the column
    if company_name_column and str(company_name_column).strip():
        return company_name_column.strip()
    
    # Second priority: Extract from LinkedIn URL
    if not linkedin_url:
        return "your organization"
    
    try:
        # Extract from URL: /company/company-name/
        parts = linkedin_url.rstrip('/').split('/')
        if 'company' in parts:
            idx = parts.index('company')
            if idx + 1 < len(parts):
                name = parts[idx + 1].replace('-', ' ').title()
                return name
    except:
        pass
    
    return "your organization"


def update_email_status(row_num, status):
    """Update email status in Excel"""
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb[INPUT_SHEET]
        headers = [cell.value for cell in ws[1]]
        
        status_idx = headers.index('email_status') + 1  # +1 for Excel 1-based indexing
        
        ws.cell(row_num, status_idx, status)
        wb.save(EXCEL_FILE)
        wb.close()
        
    except Exception as e:
        print(f"      ⚠️ Failed to update Excel: {e}")

# ==================== MAIN SENDING LOGIC ====================

def main():
    """Main execution"""

    print("\n" + "=" * 70)
    print("📧 COLD EMAIL AUTOMATION - PRODUCTION READY")
    print("=" * 70)
    print("Configuration:")
    print(f"  ✅ Emails per day: {BATCH_CONFIG['emails_per_day']}")
    print(f"  ✅ Delay between emails: {BATCH_CONFIG['delay_min']}-{BATCH_CONFIG['delay_max']}s")
    print(f"  ✅ Batch size: {BATCH_CONFIG['batch_size']} (then 5min break)")
    print(f"  ✅ Template: Static (no AI)")
    print(f"  ✅ Resume: {os.path.basename(RESUME_PATH)}")
    print("=" * 70 + "\n")

    # Check resume
    if not os.path.exists(RESUME_PATH):
        print(f"❌ ERROR: Resume not found at: {RESUME_PATH}")
        return

    # Authenticate Gmail
    print("🔐 Authenticating with Gmail API...")
    try:
        gmail_service = get_gmail_service()
        print("✅ Gmail authenticated!\n")
    except Exception as e:
        print(f"❌ Gmail authentication failed: {e}")
        return

    # Read contacts
    contacts = read_and_filter_contacts()
    
    if not contacts:
        print("✅ No pending emails to send. All done!")
        return

    # Limit to daily max
    contacts_to_send = contacts[:BATCH_CONFIG['emails_per_day']]
    remaining = len(contacts) - len(contacts_to_send)

    print(f"{'=' * 70}")
    print(f"📊 TODAY'S BATCH")
    print(f"{'=' * 70}")
    print(f"Sending today: {len(contacts_to_send)}")
    print(f"Remaining for future: {remaining}")
    print(f"{'=' * 70}\n")

    # Confirm
    proceed = input("Ready to start sending? (yes/no): ").lower().strip()
    if proceed != 'yes':
        print("   Cancelled.")
        return

    # Start sending
    print(f"\n{'=' * 70}")
    print(f"🚀 SENDING EMAILS - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"{'=' * 70}\n")

    sent_count = 0
    failed_count = 0
    start_time = time.time()

    try:
        for i, contact in enumerate(contacts_to_send):
            email = contact['email']
            first_name = contact['first_name']
            company_name = contact['company_name']
            row_num = contact['row_num']

            print(f"[{i + 1}/{len(contacts_to_send)}] {email} ({company_name})")

            # Create email body (static template)
            email_body = EMAIL_BODY_TEMPLATE.format(
                first_name=first_name,
                company_name=company_name
            )

            # Format subject with company name
            email_subject = EMAIL_SUBJECT.format(company_name=company_name)

            # Send email
            print(f"   📧 Sending email...")
            result = send_email(
                gmail_service,
                email,
                email_subject,
                email_body,
                attachment_path=None  # ← no resume for cold email
            )

            if result['success']:
                print(f"   ✅ Sent successfully!")
                sent_count += 1
                update_email_status(row_num, "Sent")
            else:
                print(f"   ❌ Failed: {result.get('error', 'Unknown error')}")
                failed_count += 1
                update_email_status(row_num, "Failed")

            # Batch break (every 10 emails)
            if (i + 1) % BATCH_CONFIG['batch_size'] == 0 and (i + 1) < len(contacts_to_send):
                print(f"\n⏸️  Batch break - waiting {BATCH_CONFIG['batch_break']//60} minutes...\n")
                time.sleep(BATCH_CONFIG['batch_break'])

            # Delay before next email (random 20-35 seconds)
            elif (i + 1) < len(contacts_to_send):
                delay = random.randint(BATCH_CONFIG['delay_min'], BATCH_CONFIG['delay_max'])
                print(f"   ⏱️  Waiting {delay}s...\n")
                time.sleep(delay)

    except KeyboardInterrupt:
        print(f"\n\n⏸️  Process interrupted!")
        print(f"   Sent: {sent_count}, Failed: {failed_count}")
        print(f"   Progress saved in Excel\n")
        return

    except Exception as e:
        print(f"\n❌ Unexpected error: {e}")
        import traceback
        traceback.print_exc()

    # Final summary
    elapsed = time.time() - start_time
    minutes = int(elapsed // 60)
    seconds = int(elapsed % 60)

    print(f"\n{'=' * 70}")
    print("🎉 BATCH COMPLETED!")
    print(f"{'=' * 70}")
    print(f"✅ Sent: {sent_count}")
    print(f"❌ Failed: {failed_count}")
    print(f"⏱️  Time: {minutes}m {seconds}s")
    print(f"📊 Remaining: {remaining}")
    print(f"{'=' * 70}\n")

    if remaining > 0:
        print(f"💡 TIP: Run again tomorrow to send next {min(remaining, BATCH_CONFIG['emails_per_day'])} emails\n")


if __name__ == "__main__":
    main()